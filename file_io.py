"""Read tabular data from CSV, TXT, and Excel input files."""

from __future__ import annotations

import csv
from collections.abc import Iterator
from pathlib import Path

SUPPORTED_INPUT_EXTENSIONS = {".csv", ".txt", ".xlsx", ".xls"}
EXCEL_EXTENSIONS = {".xlsx", ".xls"}
MAX_EXCEL_FILE_BYTES = 25 * 1024 * 1024
MAX_EXCEL_ROWS = 100_000

SUPPORTED_INPUT_LABEL = "CSV, TXT, or Excel (.xlsx, .xls)"


def input_extension(path: Path) -> str:
    return path.suffix.lower()


def is_supported_input(path: Path) -> bool:
    return input_extension(path) in SUPPORTED_INPUT_EXTENSIONS


def is_excel_input(path: Path) -> bool:
    return input_extension(path) in EXCEL_EXTENSIONS


def validate_input_file(path: Path) -> None:
    if not path.is_file():
        raise FileNotFoundError(f"File not found: {path}")
    if not is_supported_input(path):
        supported = ", ".join(sorted(SUPPORTED_INPUT_EXTENSIONS))
        raise ValueError(
            f"Unsupported file type '{path.suffix}'. Supported input types: {supported}."
        )
    if is_excel_input(path):
        size = path.stat().st_size
        if size > MAX_EXCEL_FILE_BYTES:
            max_mb = MAX_EXCEL_FILE_BYTES // (1024 * 1024)
            raise ValueError(
                f"Excel file is too large ({size / (1024 * 1024):.1f} MB). "
                f"Excel supports up to {max_mb} MB — use CSV or TXT for large files."
            )


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _validate_excel_row_count(row_count: int, path: Path) -> None:
    if row_count > MAX_EXCEL_ROWS:
        raise ValueError(
            f"Excel file '{path.name}' has {row_count:,} data rows. "
            f"Excel supports up to {MAX_EXCEL_ROWS:,} rows — use CSV or TXT for large files."
        )


def read_header(path: Path) -> list[str]:
    validate_input_file(path)
    ext = input_extension(path)

    if ext == ".csv":
        with path.open("r", encoding="utf-8", errors="replace", newline="") as infile:
            return next(csv.reader(infile))

    if ext == ".txt":
        return _read_delimited_header(path)

    if ext == ".xlsx":
        return _read_xlsx_header(path)

    if ext == ".xls":
        return _read_xls_header(path)

    raise ValueError(f"Unsupported file type: {path.suffix}")


def _read_delimited_header(path: Path) -> list[str]:
    with path.open("r", encoding="utf-8", errors="replace", newline="") as infile:
        sample = infile.read(8192)
        infile.seek(0)
        if not sample.strip():
            raise ValueError(f"File is empty: {path.name}")

        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
            reader = csv.reader(infile, dialect)
        except csv.Error:
            first_line = sample.splitlines()[0]
            if "\t" in first_line:
                reader = csv.reader(infile, delimiter="\t")
            elif "|" in first_line:
                reader = csv.reader(infile, delimiter="|")
            elif ";" in first_line and first_line.count(";") >= first_line.count(","):
                reader = csv.reader(infile, delimiter=";")
            else:
                reader = csv.reader(infile)

        header = next(reader)
        if not header or all(not cell.strip() for cell in header):
            raise ValueError(f"Could not read column headers from {path.name}")
        return [cell.strip() for cell in header]


def _read_xlsx_header(path: Path) -> list[str]:
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        if sheet is None:
            raise ValueError(f"No worksheet found in {path.name}")

        data_rows = max(sheet.max_row - 1, 0)
        _validate_excel_row_count(data_rows, path)

        row_iter = sheet.iter_rows(min_row=1, max_row=1, values_only=True)
        first_row = next(row_iter, None)
        if not first_row:
            raise ValueError(f"Excel file is empty: {path.name}")

        header = [_cell_text(value) for value in first_row]
        if not any(header):
            raise ValueError(f"Could not read column headers from {path.name}")
        return header
    finally:
        workbook.close()


def _read_xls_header(path: Path) -> list[str]:
    import xlrd

    book = xlrd.open_workbook(path)
    sheet = book.sheet_by_index(0)
    if sheet.nrows == 0:
        raise ValueError(f"Excel file is empty: {path.name}")

    _validate_excel_row_count(max(sheet.nrows - 1, 0), path)

    header = [_cell_text(sheet.cell_value(0, col_idx)) for col_idx in range(sheet.ncols)]
    if not any(header):
        raise ValueError(f"Could not read column headers from {path.name}")
    return header


def iter_data_rows(path: Path) -> Iterator[list[str]]:
    validate_input_file(path)
    ext = input_extension(path)

    if ext == ".csv":
        yield from _iter_csv_rows(path)
        return

    if ext == ".txt":
        yield from _iter_txt_rows(path)
        return

    if ext == ".xlsx":
        yield from _iter_xlsx_rows(path)
        return

    if ext == ".xls":
        yield from _iter_xls_rows(path)
        return

    raise ValueError(f"Unsupported file type: {path.suffix}")


def _iter_csv_rows(path: Path) -> Iterator[list[str]]:
    with path.open("r", encoding="utf-8", errors="replace", newline="") as infile:
        reader = csv.reader(infile)
        next(reader, None)
        for row in reader:
            yield row


def _iter_txt_rows(path: Path) -> Iterator[list[str]]:
    with path.open("r", encoding="utf-8", errors="replace", newline="") as infile:
        sample = infile.read(8192)
        infile.seek(0)

        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
            reader = csv.reader(infile, dialect)
        except csv.Error:
            first_line = sample.splitlines()[0] if sample.splitlines() else ""
            if "\t" in first_line:
                reader = csv.reader(infile, delimiter="\t")
            elif "|" in first_line:
                reader = csv.reader(infile, delimiter="|")
            elif ";" in first_line and first_line.count(";") >= first_line.count(","):
                reader = csv.reader(infile, delimiter=";")
            else:
                reader = csv.reader(infile)

        next(reader, None)
        for row in reader:
            yield row


def _iter_xlsx_rows(path: Path) -> Iterator[list[str]]:
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        if sheet is None:
            return

        row_iter = sheet.iter_rows(min_row=2, values_only=True)
        for row in row_iter:
            yield [_cell_text(value) for value in row]
    finally:
        workbook.close()


def _iter_xls_rows(path: Path) -> Iterator[list[str]]:
    import xlrd

    book = xlrd.open_workbook(path)
    sheet = book.sheet_by_index(0)
    for row_idx in range(1, sheet.nrows):
        yield [_cell_text(sheet.cell_value(row_idx, col_idx)) for col_idx in range(sheet.ncols)]


def file_type_label(path: Path) -> str:
    ext = input_extension(path)
    if ext == ".csv":
        return "CSV"
    if ext == ".txt":
        return "TXT"
    if ext in EXCEL_EXTENSIONS:
        return "Excel"
    return ext.lstrip(".").upper()
